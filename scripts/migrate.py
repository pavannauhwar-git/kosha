#!/usr/bin/env python3
"""
Kosha Migration Script
Reads your Excel budget files and imports all transactions into Supabase.

SETUP:
  1. Place your Excel files (.xlsx) in this scripts/ folder
  2. Make sure your .env file has ALL FOUR keys (see .env.example)
  3. Run:
       pip install openpyxl requests python-dotenv
       python migrate.py --dry-run     ← preview without uploading
       python migrate.py               ← import to Supabase

WHAT IT DOES:
  - Reads every monthly sheet from each Excel file
  - Assigns the 1st of the month as the date (since Excel sheets have no
    individual row dates — only a month-level label)
  - Auto-categorises transactions by keyword matching
  - Flags repayment income (Ajay EMIs, Panda EMIs) separately from salary
  - Skips carry-forward "Leftover" rows and totals/summary rows
  - Detects income in Col C+D, expenses in Col E+F, investments in Col G+H
  - Inserts in batches of 100 rows
"""

import os, sys, json, re, argparse
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    sys.exit("❌ Missing dependency. Run: pip install openpyxl")

# ── Config ─────────────────────────────────────────────────────────────────
# Default file names — override with --files argument if yours differ
EXCEL_FILES = [
    "Pavan_Budget_2023-24.xlsx",
    "Pavan_Budget_2024-25.xlsx",
    "Pavan_Budget_2025-26.xlsx",
]
BATCH_SIZE = 100

# ── Category keyword rules ──────────────────────────────────────────────────
CATEGORY_RULES = {
    'vehicle':       ['car','petrol','fuel','parking','fastag','skoda','kushaq',
                      'pnb car','rc renewal','service center','tyre'],
    'food':          ['zomato','swiggy','burger','pizza','restaurant','cafe',
                      'coffee','dunkin','kfc','mcdonalds','dominos','mcd','starbucks'],
    'groceries':     ['blinkit','zepto','bigbasket','dmart','grocer','vegetable',
                      'fruits','milk','bread','grocery'],
    'electronics':   ['samsung','iphone','apple','laptop','phone','tv','monitor',
                      'lg','sony','mi','realme','oneplus','jbl','headphone','airpods'],
    'medical':       ['pharma','hospital','doctor','clinic','medicine','apollo',
                      'health','medplus','diagnostic','lab test','heal & glow',
                      'heal and glow'],
    'travel':        ['flight','hotel','oyo','makemytrip','booking','train',
                      'irctc','bus','holiday','trip','dhanbad','airtel wifi'],
    'entertainment': ['netflix','hotstar','prime','spotify','youtube','game',
                      'steam','ps5','cinema','pvr','itunes'],
    'utilities':     ['airtel','jio','bsnl','electricity','water','gas',
                      'internet','broadband','wifi','maintenance','house maintenance'],
    'insurance':     ['lic','term plan','health insurance','car insurance','policy'],
    'credit_card':   ['icici cc','hdfc cc','credit card','cc bill','cc payment',
                      'apple wallet'],
    'shopping':      ['amazon','flipkart','myntra','ajio','nykaa','meesho'],
    'education':     ['coursera','udemy','book','course','school','tuition'],
    'personal':      ['haircut','salon','spa','gym','fitness','grooming'],
    'taxes':         ['tds','income tax','advance tax','gst'],
    'rent':          ['rent','society'],
    'gift':          ['gift','birthday','wedding','celebration'],
    'subscription':  ['subscription','apple one'],
}

REPAYMENT_KEYWORDS = [
    'loan recovery','ajay','panda -','panda payout','cbi transfer','recovery',
    'emi',   # catches "Ajay - Xth EMI", "Panda - Xth EMI"
]
SKIP_KEYWORDS = [
    'total','total cash flow','account balance','check',
    'fore-closure','foreclosure','fore closure',
]
LEFTOVER_RE = re.compile(r'leftover', re.IGNORECASE)
SAVINGS_RE  = re.compile(r'^savings$', re.IGNORECASE)

INVESTMENT_VEHICLE_MAP = {
    'esop':        'ESOPs',
    'adobe':       'Adobe ESPP',
    'ppf':         'PPF',
    'nps':         'NPS',
    'zerodha':     'Zerodha',
    'indriya':     'Indriya',
    'hsbc':        'HSBC',
    'gold':        'Gold',
    'sgb':         'SGB',
    'term plan':   'Term Plan',
    'cbi':         'CBI',
    'mutual fund': 'Mutual Fund',
}

# ── Helpers ─────────────────────────────────────────────────────────────────
def categorise(desc: str) -> str:
    dl = desc.lower()
    for cat, kws in CATEGORY_RULES.items():
        if any(kw in dl for kw in kws):
            return cat
    return 'other'

def get_vehicle(desc: str) -> str:
    dl = desc.lower()
    for kw, vehicle in INVESTMENT_VEHICLE_MAP.items():
        if kw in dl:
            return vehicle
    return 'Other'

def parse_month_sheet(name: str):
    """
    Return (month_int, year_int) from a sheet name like 'April 2023'.
    Falls back to (1, 2024) if the name is unrecognised.
    """
    MONTHS = {
        'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
        'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
    }
    parts = name.strip().lower().split()
    if len(parts) >= 2:
        m = MONTHS.get(parts[0])
        try:
            y = int(parts[-1])
            if m and 2000 <= y <= 2100:
                return m, y
        except ValueError:
            pass
    return None, None  # Signal: not a valid month sheet

def safe_float(v) -> float | None:
    """Convert a cell value to float, stripping currency symbols and commas."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v > 0 else None
    s = str(v).replace(',', '').replace('₹', '').replace('$', '').strip()
    try:
        f = float(s)
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None

def cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ''

def should_skip(inc_desc: str, exp_desc: str, inv_desc: str) -> bool:
    all_desc = (inc_desc + ' ' + exp_desc + ' ' + inv_desc).lower()
    return any(kw in all_desc for kw in SKIP_KEYWORDS)

# ── Main parser ─────────────────────────────────────────────────────────────
def parse_file(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    transactions = []

    for sheet_name in wb.sheetnames:
        month, year = parse_month_sheet(sheet_name)
        if month is None:
            # Skip non-month sheets (e.g. "Summary", "Charts")
            print(f"  ⏭  Skipping sheet: {sheet_name} (not a month sheet)")
            continue

        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name} ({month}/{year})")

        # ── FIX: Date comes from sheet name, not from column B ──────────────
        # All historical transactions in a sheet share the 1st of that month.
        # Column B in your Excel contains the month label (row 2) and is
        # otherwise empty — so we NEVER use it as a date source.
        sheet_date = f"{year}-{month:02d}-01"

        sheet_count = 0
        for row in range(2, ws.max_row + 1):
            # ── FIX: Detect rows by whether Col C, E, or G has a numeric value ──
            # (not by Col B which is empty for all data rows)
            inc_amt  = safe_float(ws.cell(row=row, column=3).value)  # C
            inc_desc = cell_str(ws, row, 4)                           # D
            exp_amt  = safe_float(ws.cell(row=row, column=5).value)  # E
            exp_desc = cell_str(ws, row, 6)                           # F
            inv_amt  = safe_float(ws.cell(row=row, column=7).value)  # G
            inv_desc = cell_str(ws, row, 8)                           # H

            # Skip rows with no numeric data at all
            if not inc_amt and not exp_amt and not inv_amt:
                continue

            # Skip total/summary rows
            if should_skip(inc_desc, exp_desc, inv_desc):
                continue

            # ── Income ──────────────────────────────────────────────────────
            if inc_amt and inc_desc:
                # Skip "Leftover (Month Year)" carry-forward rows
                if LEFTOVER_RE.search(inc_desc):
                    continue
                # Skip bare "Savings" row (internal transfer)
                if SAVINGS_RE.match(inc_desc):
                    continue
                desc_l    = inc_desc.lower()
                is_repay  = any(kw in desc_l for kw in REPAYMENT_KEYWORDS)
                transactions.append({
                    'date':              sheet_date,
                    'type':              'income',
                    'description':       inc_desc,
                    'amount':            inc_amt,
                    'category':          'transfer' if is_repay else categorise(inc_desc),
                    'is_repayment':      is_repay,
                    'payment_mode':      'net_banking',
                    'investment_vehicle': None,
                    'notes':             None,
                })
                sheet_count += 1

            # ── Expense ─────────────────────────────────────────────────────
            if exp_amt and exp_desc:
                desc_l = exp_desc.lower()
                if any(kw in desc_l for kw in SKIP_KEYWORDS):
                    continue
                transactions.append({
                    'date':              sheet_date,
                    'type':              'expense',
                    'description':       exp_desc,
                    'amount':            exp_amt,
                    'category':          categorise(exp_desc),
                    'is_repayment':      False,
                    'payment_mode':      'upi',
                    'investment_vehicle': None,
                    'notes':             None,
                })
                sheet_count += 1

            # ── Investment ───────────────────────────────────────────────────
            if inv_amt and inv_desc:
                transactions.append({
                    'date':              sheet_date,
                    'type':              'investment',
                    'description':       inv_desc,
                    'amount':            inv_amt,
                    'category':          'other',
                    'is_repayment':      False,
                    'payment_mode':      'net_banking',
                    'investment_vehicle': get_vehicle(inv_desc),
                    'notes':             None,
                })
                sheet_count += 1

        print(f"     → {sheet_count} transactions")

    return transactions

# ── Supabase upload ─────────────────────────────────────────────────────────
def upload(transactions: list[dict], url: str, key: str) -> int:
    import urllib.request
    headers = {
        'apikey':        key,
        'Authorization': f'Bearer {key}',
        'Content-Type':  'application/json',
        'Prefer':        'return=minimal',
    }
    endpoint = f"{url.rstrip('/')}/rest/v1/transactions"
    total    = 0

    for i in range(0, len(transactions), BATCH_SIZE):
        batch = transactions[i:i + BATCH_SIZE]
        data  = json.dumps(batch, default=str).encode()
        req   = urllib.request.Request(endpoint, data=data, headers=headers, method='POST')
        try:
            with urllib.request.urlopen(req):
                total += len(batch)
                print(f"  ✅ Batch {i//BATCH_SIZE + 1}: {len(batch)} rows (total {total})")
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            print(f"  ❌ Batch {i//BATCH_SIZE + 1} failed [{e.code}]: {body[:200]}")

    return total

# ── Entry point ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Kosha — Excel → Supabase migration')
    parser.add_argument('--dry-run', action='store_true',
                        help='Parse files and preview results without uploading')
    parser.add_argument('--files', nargs='+', default=EXCEL_FILES,
                        help='Excel file paths (default: looks for the 3 Pavan_Budget files)')
    args = parser.parse_args()

    # ── FIX: Load env — look for SUPABASE_URL and SUPABASE_KEY (no VITE_ prefix)
    # The VITE_ prefix is only needed by the React app (Vite's browser env system).
    # This Python script reads the .env file directly and needs the plain names.
    supabase_url = os.environ.get('SUPABASE_URL')
    supabase_key = os.environ.get('SUPABASE_KEY')

    env_file = Path('.env')
    if not env_file.exists():
        # Try one level up (in case script is run from scripts/ subfolder)
        env_file = Path('../.env')

    if env_file.exists():
        for line in env_file.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            k, _, v = line.partition('=')
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            if k == 'SUPABASE_URL'  and not supabase_url: supabase_url = v
            if k == 'SUPABASE_KEY'  and not supabase_key: supabase_key = v
    else:
        print("⚠️  No .env file found. Reading from environment variables only.")

    # ── Parse all files ──────────────────────────────────────────────────────
    all_txns = []
    for f in args.files:
        p = Path(f)
        if not p.exists():
            # Also try relative to script location
            alt = Path(__file__).parent / f
            if alt.exists():
                p = alt
            else:
                print(f"⚠️  File not found: {f} — skipping")
                continue
        print(f"\n📂 Parsing: {p.name}")
        txns = parse_file(str(p))
        all_txns.extend(txns)

    total_found = len(all_txns)
    print(f"\n📊 Total: {total_found} transactions across {len(args.files)} file(s)")

    if total_found == 0:
        print("\n⚠️  No transactions found.")
        print("   Check that your Excel files are in the scripts/ folder")
        print("   and that transaction data is in columns C–H.")
        return

    # ── Dry run ──────────────────────────────────────────────────────────────
    if args.dry_run:
        out = Path('migrated_transactions.json')
        out.write_text(json.dumps(all_txns, indent=2, default=str), encoding='utf-8')
        print(f"\n✅ Dry run complete — saved to {out}")
        print("\nSample (first 5 transactions):")
        print(f"  {'Date':<12} {'Type':<12} {'Amount':>10}  Description")
        print(f"  {'-'*12} {'-'*12} {'-'*10}  {'-'*30}")
        for t in all_txns[:5]:
            print(f"  {t['date']:<12} {t['type']:<12} {t['amount']:>10,.0f}  {t['description'][:40]}")
        return

    # ── Upload ───────────────────────────────────────────────────────────────
    if not supabase_url or not supabase_key:
        print("\n❌ SUPABASE_URL and SUPABASE_KEY must be set in your .env file.")
        print("   Open .env and make sure these two lines exist (no VITE_ prefix):")
        print("     SUPABASE_URL=https://your-project.supabase.co")
        print("     SUPABASE_KEY=your-service-role-secret-key")
        print("\n   Find your service_role key in:")
        print("   Supabase Dashboard → Settings → API Keys → service_role / Secret")
        sys.exit(1)

    print(f"\n🚀 Uploading to {supabase_url}…")
    n = upload(all_txns, supabase_url, supabase_key)
    print(f"\n✅ Migration complete — {n} of {total_found} transactions uploaded")

if __name__ == '__main__':
    main()
